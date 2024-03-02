import uuid, os, json, re
from datetime import datetime

from dotenv import load_dotenv
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import numbers

from file_operations import FileOperations


# logger.add("debug.log", format='{time} | {level} | {file} | {name} | {function} | {line} | {message}', level='DEBUG', rotation='1 week', compression='zip')

fo = FileOperations()

def create_extra_id() -> str:
    ''' generate random UUID fot extra_id '''
    return str(uuid.uuid4())


def get_request_params_minus_messages(path_file=None,
                                      data=None,
                                      request_params=None,
                                      double=False) -> dict:
        ''' this method removes from request_params any data in format:
        {
            unp: {
                "payment_date": "%d.%m.%Y",
                any...
            }
        }
        Define double=True for saving in request_params recicpent
        if payment_date of request_params > payment_date of FIRST_SUCCESS_MESSAGES '''
        logger.info("Start 'utils.get_request_params_minus_messages' method")
        if not request_params:
            request_params = fo.get_last_element(path_folder=os.getenv("VIRGIN_REQ_PAR_MASS_BROAD"))
        if path_file:
            data = fo.get_data_from_json_file(path_file=path_file)
        recipients = []
        for recipient in request_params["recipients"]:
            str_unp = str(recipient["unp"])
            if str_unp not in data:
                recipients.append(recipient)
            else:
                rq_pay_date = datetime.strptime(recipient["payment_date"], "%d.%m.%Y").date()
                data_pay_date = datetime.strptime(data[str_unp]["payment_date"], "%d.%m.%Y").date()
                if double and rq_pay_date > data_pay_date:
                    recipients.append(recipient)
        request_params["recipients"] = recipients
        logger.info("End 'utils.get_request_params_minus_messages' method")
        return request_params


def remove_message_from_success(unp=None, unp_list=None, double=None) -> None:
    success_file_name =  "SAVE_FILE_SUCCESS_MESSAGES_DOUBLE" if double else "SAVE_FILE_SUCCESS_MESSAGES_FIRST"
    success_messages = fo.get_data_from_json_file(os.getenv(success_file_name))
    if unp:
        if unp in success_messages:
            del success_messages[unp]
    if unp_list:
        for unp in unp_list:
            if unp in success_messages:
                del success_messages[unp]
    fo.save_file(data_list=success_messages, full_file_name=os.getenv(success_file_name))


class CreateUpdateInvalidXlsFile:

    def check_if_file_exist(self, path:str) -> bool:
        ''' метод проверяет, существует ли файл '''
        return os.path.isfile(path)

    def create_file_if_not_exist(self, path:str, sheet_name:str, headers:list) -> None:
        ''' метод создает файл с расширением xlsx, присваивает имя первому листу,
        вносит заголовки таблицы и применяет к ним форматирование '''
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        ft = Font(bold=True)
        len_coll = ws.max_column
        for coll in ws.iter_cols(max_row=1):
            for cell in coll:
                cell.font = ft
        wb.save(path)

    def update_file_if_exist(self, path:str, insert_data:list) -> None:
        ''' метод добавляет записи в файл с раширением xlsx '''
        wb = load_workbook(filename=path)
        ws = wb.active
        for row in insert_data:
            ws.append(row)
        wb.save(path)

    def get_list_of_column(self, path:str, col:str) -> list:
        ''' метод возвращает лист значений, содержащихся в ячейках переданного столбца '''
        wb = load_workbook(filename=path)
        ws = wb.active
        column_list = []
        for cell in ws[col]:
            column_list.append(cell.value)
        return column_list

    def get_unps_not_exist_in_file(self, path:str, col:str, unp_list:list) -> list:
        ''' метод возвращает список УНП, не содержащихся в переданном столбце переданного файла '''
        exist_unps = self.get_list_of_column(path=path, col=col)
        unps_not_exist_in_file = []
        for unp in unp_list:
            if unp not in exist_unps:
                unps_not_exist_in_file.append(unp)
        return unps_not_exist_in_file

    def get_data_by_unps(self, unp_list:list, data:dict) -> dict:
        ''' метод возвращает список из словарей по совпадающим ключам, переданным в unp_list '''
        new_data = {}
        for unp, company_data in data.items():
            if unp in unp_list:
                new_data[unp] = company_data
        return new_data

    def transform_insert_data_from_dict(self, data, *args) -> list:
        ''' метод преобразует словарь в список из списков с данными, найденными в словаре по ключам, переданным в args '''
        valid_data = []
        for unp, company_data in data.items():
            if unp == "None":
                continue
            internal_list = [int(unp)]
            for key in args:
                if key == "phone_number":
                    internal_list.append(company_data[str(key)])
                else:
                    internal_list.append(company_data[key])
            copy_internal_list = internal_list.copy()
            valid_data.append(copy_internal_list)
        return valid_data
    
    def set_collumn_string_format(self, path:str, coll:int) -> None:
        wb = load_workbook(filename=path)
        ws = wb.active
        for row in ws[2:ws.max_row]:
            cell = row[coll]
            cell.number_format = numbers.FORMAT_NUMBER
        wb.save(path)

    def create_update_file(self, path_from:str, path_to:str, sheet_name:str) -> None:
        ''' метод определяет порядок вызова функций для создания и обновления файла '''
        invalid_phones = fo.get_data_from_json_file(path_file=path_from)
        unp_list_invalid_phones = list(invalid_phones.keys())
        if not self.check_if_file_exist(path=path_to):
            self.create_file_if_not_exist(path=path_to, sheet_name=sheet_name,
                                          headers=["УНП", "ИМЯ", "ТЕЛЕФОН", "ДАТА ВОЗН. ЗАДОЛЖЕННОСТИ"])
        unps_not_exist_in_file = self.get_unps_not_exist_in_file(path=path_to, col="A", unp_list=unp_list_invalid_phones)
        insert_data = self.get_data_by_unps(unp_list=unps_not_exist_in_file, data=invalid_phones)
        transform_data = self.transform_insert_data_from_dict(insert_data, "company_name", "phone_number", "payment_date")
        self.update_file_if_exist(path=path_to, insert_data=transform_data)
        self.set_collumn_string_format(path=path_to, coll=2)
        self.set_collumn_string_format(path=path_to, coll=0)

    def create_update_invalid_phones(self, path_from=os.getenv("PATH_UNCORRECT_PHONE_NUMBERS"),
                                     path_to=os.getenv("INVALID_PHONES"), sheet_name="invalid_phones") -> None:
        ''' метод вызывает create_update_file, передавая данные для некорректных номеров '''
        logger.info("Start 'utils.CreateUpdateInvalidXlsFile.create_update_invalid_phones'")
        self.create_update_file(path_from=path_from, path_to=path_to, sheet_name=sheet_name)
        logger.info("End 'utils.CreateUpdateInvalidXlsFile.create_update_invalid_phones'")

    def create_update_error_phones(self, path_from=os.getenv("FILE_FAIL_MESSAGES"),
                                   path_to=os.getenv("ERROR_XLS_PHONES"), sheet_name="error_phones") -> None:
        ''' метод вызывает create_update_file, передавая данные для номеров, отправка на которые завершилась с ошибкой '''
        logger.info("Start 'utils.CreateUpdateInvalidXlsFile.create_update_error_phones'")
        self.create_update_file(path_from=path_from, path_to=path_to, sheet_name=sheet_name)
        logger.info("Start 'utils.CreateUpdateInvalidXlsFile.create_update_error_phones'")


@logger.catch
def main() -> None:
    invphones = CreateUpdateInvalidXlsFile()
    invphones.create_update_invalid_phones()

if __name__ == "__main__":
    main()
    
