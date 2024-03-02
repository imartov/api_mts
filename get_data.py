import json, os
from datetime import datetime, timedelta

from dotenv import load_dotenv
from openpyxl import load_workbook
from loguru import logger

from createrp import MassBroadcast, OneMessage
from phone import PhoneOperations
from file_operations import FileOperations, report_message_form
from utils import *
from checking import CheckReportJobId
from analysis import Analysis


# logger.add("debug.log", format='{time} | {level} | {file} | {name} | {function} | {line} | {message}', level='DEBUG', rotation='1 week', compression='zip')

report_message = {
    "all_count_clients": 0,
    "full_debt_sum": 0,
    "debtor_count_after_filter": 0,
    "debtor_count_valid_number": 0,
    "debtor_count_unvalid_number": 0,
    "debt_sum_valid_number": 0,
    "debt_sum_unvalid_number": 0
}

fo = FileOperations()
load_dotenv()

class GetData:
    def __init__(self, mass_broadcast=True, one_message=False) -> None:
        if mass_broadcast:
            self.rp = MassBroadcast()
            del self.rp.request_params["recipients"][0]
        elif one_message:
            self.rp = OneMessage()

    def parse_xl(self) -> dict:
        ''' open excel file and parse it '''
        logger.info("Start 'get_data.FileOperations.parse_xl' method")
        wb = load_workbook(filename=os.getenv("EXCEL_FILE"))
        # wb = load_workbook(filename=os.getenv("EXCEL_FILE_PROD"))
        ws = wb.active
        checking = CheckReportJobId()
        analysis = Analysis()
        removestoplist = RemoveStopList()
        cfilter = Filter()

        result_sum = 0
        phone_operations = PhoneOperations()
        for row in ws.iter_rows(min_row=3):
            unp = row[11]
            company_name = row[1]
            debt_sum = row[4]
            phone_number = row[10]
            payment_date = row[12]

            # end of dataframe
            if not unp.value and not company_name.value and not payment_date.value:
                break

            report_message["all_count_clients"] += 1
            
            # check if unp exist
            if not unp.value:
                continue
            
            # check if debt sum exists
            if not debt_sum.value:
                # remove from success messages if company doesn't have debt_sum
                remove_message_from_success(unp=str(unp.value))
                remove_message_from_success(unp=str(unp.value), double=True) 
                continue

            report_message["full_debt_sum"] += round(int(debt_sum.value), 2)
            
            # check if payment day is earlier than define in Filter.start_date var
            if not cfilter.date_filter(date=payment_date.value):
                continue

            # check if debt sum is less than define in Filter.min_sum var
            if not cfilter.sum_filter(sum=debt_sum.value):
                result_sum += int(debt_sum.value)
                continue
            
            # checking whether the debt arose in a certain period
            if not check_debt_period(payment_date=payment_date.value):
                continue

            # check if unp in stop-list
            if removestoplist.check_if_in_lists(unp=unp.value):
                continue
            
            report_message["debtor_count_after_filter"] += 1

            # check valid phone number
            valid_phone_number = phone_operations.check_phone_number(
            unp=str(unp.value), company_data={
                "company_name": company_name.value,
                "payment_date": payment_date.value,
                "phone_number": phone_number.value
                }
            )
            if valid_phone_number:
                report_message["debtor_count_valid_number"] += 1
                report_message["debt_sum_valid_number"] += round(int(debt_sum.value), 2)
                exist_first = checking.check_exist_success_message(
                    unp=int(unp.value),
                    payment_date=payment_date.value
                )
                exist_double = checking.check_exist_success_message(
                    unp=int(unp.value),
                    payment_date=payment_date.value, double=True
                )
                # pass debtors to virgin request params
                if not exist_first and not exist_double:
                    self.rp.create(
                        phone_number=valid_phone_number,
                        company_name=company_name.value,
                        debt_sum=debt_sum.value,
                        unp=int(unp.value),
                        payment_date=payment_date.value
                    )
            else:
                report_message["debtor_count_unvalid_number"] += 1
                report_message["debt_sum_unvalid_number"] += round(int(debt_sum.value), 2)
                    
        copy_request_params = dict(self.rp.request_params)
        fo.save_data(data=copy_request_params,
                     path_to_folder=os.getenv("VIRGIN_REQ_PAR_MASS_BROAD"))
        report_message_form(labels=report_message)
        logger.info("End 'get_data.FileOperations.parse_xl' method")
        print("СУмма: ", result_sum)
        return self.rp.request_params


class RemoveStopList:
    ''' класс проверяет, находится ли унп должника в списках из УНП '''
    def __init__(self) -> None:
        self.unp_stop_list = self.get_unp_stop_list()
        self.unp_phone_list = self.get_unp_phone_list()
        self.unp_fail_messages_list = self.get_unp_fail_messages_list()

    def get_unp_stop_list(self) -> list:
        wb = load_workbook(filename=os.getenv("STOP_CLIENTS_LIST"))
        ws = wb.active
        unp_stop_list = []
        for row in ws.iter_rows(min_col=2, max_col=6, min_row=4):
            unp_stop_list.append(int(row[0].value))
        return unp_stop_list
    
    def get_unp_phone_list(self) -> list:
        wb = load_workbook(filename=os.getenv("STOP_PHONE_LIST"))
        ws = wb.active
        unp_phone_list = []
        for row in ws.iter_rows(min_col=2, max_col=6, min_row=4):
            unp_phone_list.append(int(row[1].value))
        return unp_phone_list
    
    def get_unp_fail_messages_list(self) -> list:
        fail_messages = fo.get_data_from_json_file(path_file=os.getenv("FILE_FAIL_MESSAGES"))
        unp_fail_messages_list = list(fail_messages.keys())
        return unp_fail_messages_list

    def check_if_in_lists(self, unp) -> bool:
        if int(unp) in self.unp_stop_list or int(unp) in self.unp_phone_list or str(unp) in self.unp_fail_messages_list:
            return True
        return False
    

class Filter:
    def __init__(self) -> None:
        self.min_sum = 50
        self.start_date = datetime(2021, 1, 1).date()

    def sum_filter(self, sum) -> bool:
        if int(sum) >= self.min_sum:
            return True
        return False
    
    def date_filter(self, date:str) -> bool:
        if datetime.strptime(date, "%d.%m.%Y").date() >= self.start_date:
            return True
        return False
    

class CheckSuccessIfNeedSend:
    ''' класс определяет, необходимо ли направлять сообщение должнику,
     находящемуся в success_messages '''
    def __init__(self) -> None:
        pass

    def check_payment_date_later_than_success(self, unp:str, payment_date:str, double=False) -> bool:
        ''' метод проверяет переданную дату долга клиента:
         если переданная дата долга позже, чем сохраненная в success_messages, верент True, иначе - False '''
        path_success = "SAVE_FILE_SUCCESS_MESSAGES_DOUBLE" if double else "SAVE_FILE_SUCCESS_MESSAGES_FIRST"
        success_messages = fo.get_data_from_json_file(os.getenv(path_success))
        if unp in success_messages:
            success_payment_date = datetime.strptime(success_messages[unp]["payment_date"], "%d.%m.%Y")
            if datetime.strptime(payment_date.strip(), "%d.%m.%Y") > success_payment_date:
                return True

    def check_equal_payment_date_for_double_message(self, unp:str, payment_date:str) -> bool:
        ''' метод проверяет равна ли переданная дата задолженности дате, сохраненной в first success_messages,
         если даты равны, метод вернет True, иначе - False '''
        success_messages = fo.get_data_from_json_file(os.getenv("SAVE_FILE_SUCCESS_MESSAGES_FIRST"))
        if unp in success_messages:
            success_payment_date = datetime.strptime(success_messages[unp]["payment_date"], "%d.%m.%Y")
            if datetime.strptime(payment_date.strip(), "%d.%m.%Y") == success_payment_date:
                return True

    def check_daelay_days_for_double_message(self, unp:str) -> bool:
        ''' метод проверяет дату отправки first success_mesasges для направления второго сообщения
         если дата отправки + установленное количество дней меньше либо равно текущей дате
         метод вернет True, иначе - False '''
        success_messages = fo.get_data_from_json_file(os.getenv("SAVE_FILE_SUCCESS_MESSAGES_FIRST"))
        if unp in success_messages:
            delivering_date = datetime.strptime(success_messages[unp]["delivering_date"], fo.strftime_datatime_format).date()
            if delivering_date + timedelta(days=int(os.getenv("DELAY_DAYS_FOR_DOUBLE"))) <= datetime.now().date():
                return True
    
    def run_for_pass_to_double(self, unp:str, payment_date:str) -> None:
        ''' метод определяет порядок выполнения методов класса для передачи
         recipients в параметры запроса второго сообщения '''
        if self.check_equal_payment_date_for_double_message(unp=unp, payment_date=payment_date):
            if self.check_daelay_days_for_double_message(unp=unp):
                return True
            
    def run_for_pass_to_first(self, unp:str, payment_date:str) -> None:
        ''' метод определяет порядок выполнения методов класса для передачи
         recipients в параметры запроса первого сообщения '''
        if self.check_payment_date_later_than_success(unp=unp, payment_date=payment_date) or\
        self.check_payment_date_later_than_success(unp=unp, payment_date=payment_date, double=True):
            return True


def check_debt_period(payment_date, count_days=2) -> bool:
    ''' метод проверяет дату образования задолженности:
     если дата образования задолженности + count_days меньше либо равна текщей дате,
     метод вернет True, иначе - None '''
    payment_date = datetime.strptime(str(payment_date), "%d.%m.%Y").date()
    if payment_date + timedelta(days=count_days) <= datetime.now().date():
        return True
    return False


def main() -> None:
    rq = GetData()
    request_params = rq.parse_xl()

if __name__ == "__main__":
    main()