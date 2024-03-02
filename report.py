import os, json

from openpyxl import Workbook, load_workbook

from utils import create_extra_id
from createrp import MassBroadcast
from api_mts import ApiMTS
from file_operations import FileOperations
from checking import CheckReportJobId


fo = FileOperations()
cr = CheckReportJobId()

class Report:
    def get_phone_list_for_report(self) -> list:
        wb = load_workbook(filename=os.getenv("REPORT_PHONES"))
        ws = wb.active
        phone_list = []
        for row in ws.iter_rows(min_row=3):
            phone_list.append(int(row[0].value))
        return phone_list

    def create_text_for_report(self) -> str:
        # path_text_file_name = "NOTICE_SMS_FAIL_TEXT" if fail else "NOTICE_SMS_SUCCESS_TEXT"
        with open(os.getenv("NOTICE_SMS_SUCCESS_TEXT"), "r", encoding="utf-8") as file:
            text = file.read()
        with open(os.getenv("JSON_REPORT_MESSAGE"), "r", encoding="utf-8") as file:
            labels = json.load(file)
        text = text.format(**labels)
        return text
    
    def create_recipients_request_params(self, phone_list:list) -> list:
        recipients = []
        for phone in phone_list:
            recipients.append(
                {
                    "phone_number": phone,
                    "extra_id": create_extra_id(),
                }
            )
        return recipients
    
    def get_request_params(self, recipients:list) -> dict:
        cr = MassBroadcast(text_message=self.create_text_for_report())
        request_params = cr.create(recipients=recipients)
        request_params["tag"] = "report_debt_collection"
        return request_params

    def run(self) -> None:
        phone_list = self.get_phone_list_for_report()
        recipients = self.create_recipients_request_params(phone_list=phone_list)
        request_params = self.get_request_params(recipients=recipients)
        fo.save_data(data=request_params, path_to_folder=os.getenv("SAVE_FIRST_REQ_PAR_MASS_BROAD"))
        am = ApiMTS()
        send_message = am.send_broadcast_mass_messages_and_get_report_by_job_id(request_params=request_params)
        if send_message:
            fo.save_data(data=send_message["resp_message"], path_to_folder=os.getenv("SAVE_RESPONSE_DATA"))
            fo.save_data(data=send_message["resp_report"], path_to_folder=os.getenv("SAVE_FIRST_REPORTS_JOB_ID"))


def main() -> None:
    rep = Report()
    rep.run()

if __name__ == "__main__":
    main()